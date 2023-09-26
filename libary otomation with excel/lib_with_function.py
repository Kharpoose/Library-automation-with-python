from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font



wb = Workbook()
ws = wb.active
ws.title = "Library"


ws.append(['Books Name', 'date of borrowing the book',
          'date of return of the book', 'the person who rented the book'])
ws['A1'].font = Font(color= 'b7312c')
ws['B1'].font = Font(color= 'b7312c')
ws['C1'].font = Font(color= 'b7312c')
ws['D1'].font = Font(color= 'b7312c')


def Library():
    aaa = True
    while aaa == True:
        book_name = input("input name a book: ")
        timein = input("the date the book was checked out: ")
        timeout = input("input the return date: ")
        name = input("Who is taking the book? ")

        ws.append([book_name, timein, timeout, name])

        b = str(input("print exit if you want to exit otherwise print anything"))

        if b == "exit":
            aaa = False
        wb.save('libraryNew.xlsx')


def editLibrary():
    ld = load_workbook('libraryNew.xlsx')
    ls = ld.active
    y = input("input ")
    print(ws[f"{y}"].value)

x = str(input("if you want to start new Library Automation write start or if you want to edit your library write edit: ")).lower()

if x == "start":
    Library()
elif x == "edit":
    editLibrary() 
