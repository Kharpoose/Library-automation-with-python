from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill



wb = Workbook()
ws = wb.active
ws.title = "Library"

ws.append(['Books Name','date of borrowing the book','date of return of the book','the person who rented the book'])
ws['A1'].font = Font(color= 'b7312c')
ws['B1'].font = Font(color= 'b7312c')
ws['C1'].font = Font(color= 'b7312c')
ws['D1'].font = Font(color= 'b7312c')

book_name = input("input name a book: ")
timein = input("the date the book was checked out: ")
timeout = input("input the return date: ")
name = input("Who is taking the book? ")

ws["A2"]  = book_name
ws["B2"]  = timein
ws["C2"]  = timeout
ws["D2"]  = name


wb.save('library.xlsx')

#lib with function is better i think