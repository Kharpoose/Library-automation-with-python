from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font



#wb = Workbook()
#ws.title = "Library"

wb = load_workbook('libraryNewnewnew.xlsx')

ws = wb.active
print(ws["A1"].value)


def editLibrary():
    ld = load_workbook('libraryNew.xlsx')
    ls = ld.active
    y = input("input ")
    print(ws[f"{y}"].value)
    
    
editLibrary()